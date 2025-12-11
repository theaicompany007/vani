"""
Batch Contact Import Script
Imports contacts from Excel file with batch processing, threading, and memory management.
Designed for large files (1000+ records).

Usage:
    python scripts/import_contacts_batch.py <excel_file_path> [options]

Options:
    --batch-size: Number of records per batch (default: 100)
    --threads: Number of parallel threads (default: 4)
    --update-existing: Update existing contacts (default: False)
    --import-only-new: Only import new contacts, skip duplicates (default: False)
    --sheet: Specific sheet name to import (default: all sheets)
    --dry-run: Preview mode, don't actually import (default: False)
"""

import sys
import os
import argparse
import logging
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Optional
import openpyxl
from io import BytesIO
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
import time

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from app.supabase_client import get_supabase_client
from app.services.contact_service import (
    upsert_contacts,
    find_duplicates,
    resolve_best_domain,
    normalize_email,
    normalize_phone
)
from flask import Flask

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(f'logs/batch_import_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class BatchContactImporter:
    """Batch contact importer with threading and memory management"""
    
    def __init__(self, excel_path: str, batch_size: int = 100, num_threads: int = 4,
                 update_existing: bool = False, import_only_new: bool = False,
                 sheet_name: Optional[str] = None, dry_run: bool = False):
        self.excel_path = excel_path
        self.batch_size = batch_size
        self.num_threads = num_threads
        self.update_existing = update_existing
        self.import_only_new = import_only_new
        self.sheet_name = sheet_name
        self.dry_run = dry_run
        
        # Create Flask app context for Supabase
        from dotenv import load_dotenv
        load_dotenv()
        load_dotenv('.env.local', override=True)
        
        self.app = Flask(__name__)
        self.app.config['SUPABASE_URL'] = os.getenv('SUPABASE_URL')
        self.app.config['SUPABASE_KEY'] = os.getenv('SUPABASE_SERVICE_KEY') or os.getenv('SUPABASE_KEY')
        self.app.config['SUPABASE_SERVICE_KEY'] = os.getenv('SUPABASE_SERVICE_KEY')
        
        # Statistics
        self.stats = {
            'total_rows': 0,
            'processed': 0,
            'imported': 0,
            'updated': 0,
            'skipped': 0,
            'errors': 0,
            'batches': 0,
            'start_time': None,
            'end_time': None
        }
        
        # Thread-safe locks
        self.stats_lock = threading.Lock()
        self.errors: List[Dict[str, Any]] = []
        self.errors_lock = threading.Lock()
    
    def load_excel_file(self) -> List[Dict[str, Any]]:
        """Load and parse Excel file"""
        logger.info(f"Loading Excel file: {self.excel_path}")
        
        if not os.path.exists(self.excel_path):
            raise FileNotFoundError(f"Excel file not found: {self.excel_path}")
        
        workbook = openpyxl.load_workbook(self.excel_path, data_only=True, read_only=True)
        
        # Select sheets to process
        if self.sheet_name:
            if self.sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{self.sheet_name}' not found in Excel file")
            sheet_names = [self.sheet_name]
        else:
            sheet_names = workbook.sheetnames
            logger.info(f"Processing {len(sheet_names)} sheet(s): {', '.join(sheet_names)}")
        
        all_contacts = []
        
        for sheet_name in sheet_names:
            logger.info(f"Processing sheet: {sheet_name}")
            sheet = workbook[sheet_name]
            
            # Get headers from first row
            headers = []
            for cell in sheet[1]:
                header = str(cell.value).strip() if cell.value else ''
                # Auto-map common column names
                header_lower = header.lower().replace(' ', '_').replace('-', '_')
                if header_lower in ['name', 'full_name', 'contact_name']:
                    header = 'name'
                elif header_lower in ['email', 'email_address', 'e_mail']:
                    header = 'email'
                elif header_lower in ['phone', 'phone_number', 'mobile', 'tel']:
                    header = 'phone'
                elif header_lower in ['company', 'company_name', 'organization']:
                    header = 'company'
                elif header_lower in ['role', 'title', 'job_title', 'position', 'designation']:
                    header = 'role'
                elif header_lower in ['linkedin', 'linkedin_url', 'linkedin_profile']:
                    header = 'linkedin'
                elif header_lower in ['industry', 'sector']:
                    header = 'industry'
                elif header_lower in ['city', 'location']:
                    header = 'city'
                elif header_lower in ['lead_source', 'source', 'source_name']:
                    header = 'lead_source'
                headers.append(header)
            
            # Parse rows
            row_count = 0
            skipped_count = 0
            for row in sheet.iter_rows(min_row=2, values_only=False):
                row_count += 1
                contact_data = {}
                for idx, cell in enumerate(row):
                    if idx < len(headers) and headers[idx] and cell.value:
                        contact_data[headers[idx]] = str(cell.value).strip()
                
                # Skip completely empty rows
                has_name = bool(contact_data.get('name'))
                has_email = bool(contact_data.get('email'))
                has_phone = bool(contact_data.get('phone'))
                
                if not has_name and not has_email and not has_phone:
                    skipped_count += 1
                    continue
                
                # Add sheet name
                contact_data['sheet'] = sheet_name
                all_contacts.append(contact_data)
            
            logger.info(f"Sheet '{sheet_name}': Processed {row_count} rows, extracted {len([c for c in all_contacts if c.get('sheet') == sheet_name])} contacts, skipped {skipped_count} empty rows")
        
        workbook.close()
        
        # Normalize rows
        normalized_rows = []
        for row in all_contacts:
            normalized = {**row}
            
            # Extract domain from email if not provided
            if not normalized.get('domain') and normalized.get('email'):
                normalized['domain'] = resolve_best_domain(None, normalized['email'])
            
            # Ensure company name exists if we have domain but no company
            if not normalized.get('company') and not normalized.get('company_name') and normalized.get('domain'):
                domain_parts = normalized['domain'].split('.')
                if domain_parts:
                    normalized['company'] = domain_parts[0].title()
            
            # Normalize lead_source
            if normalized.get('leadSource') and not normalized.get('lead_source'):
                normalized['lead_source'] = normalized['leadSource']
            if not normalized.get('lead_source') and not normalized.get('leadSource') and normalized.get('sheet'):
                normalized['lead_source'] = normalized['sheet']
            
            normalized_rows.append(normalized)
        
        self.stats['total_rows'] = len(normalized_rows)
        logger.info(f"Total contacts to process: {len(normalized_rows)}")
        
        return normalized_rows
    
    def process_batch(self, batch: List[Dict[str, Any]], batch_num: int) -> Dict[str, Any]:
        """Process a single batch of contacts"""
        with self.app.app_context():
            try:
                # Initialize Supabase client directly
                from supabase import create_client
                supabase_url = self.app.config.get('SUPABASE_URL')
                supabase_key = self.app.config.get('SUPABASE_SERVICE_KEY')
                
                if not supabase_url or not supabase_key:
                    raise Exception("Supabase configuration not available")
                
                supabase = create_client(supabase_url, supabase_key)
                if not supabase:
                    raise Exception("Supabase client not available")
                
                # Find duplicates if import_only_new
                if self.import_only_new:
                    duplicates, uniques = find_duplicates(supabase, batch)
                    rows_to_import = uniques
                    logger.info(f"Batch {batch_num}: {len(uniques)} new, {len(duplicates)} duplicates skipped")
                else:
                    rows_to_import = batch
                
                if self.dry_run:
                    logger.info(f"Batch {batch_num}: DRY RUN - Would import {len(rows_to_import)} contacts")
                    return {
                        'batch_num': batch_num,
                        'imported': len(rows_to_import),
                        'errors': []
                    }
                
                # Import batch
                result = upsert_contacts(supabase, rows_to_import, {'updateExisting': self.update_existing})
                
                imported = result.get('imported', 0)
                errors = result.get('errors', [])
                
                with self.stats_lock:
                    self.stats['imported'] += imported
                    self.stats['errors'] += len(errors)
                    self.stats['processed'] += len(batch)
                
                if errors:
                    with self.errors_lock:
                        self.errors.extend(errors)
                
                logger.info(f"Batch {batch_num}: Imported {imported}/{len(rows_to_import)} contacts")
                
                return {
                    'batch_num': batch_num,
                    'imported': imported,
                    'errors': errors
                }
                
            except Exception as e:
                logger.error(f"Batch {batch_num} error: {e}", exc_info=True)
                with self.stats_lock:
                    self.stats['errors'] += len(batch)
                with self.errors_lock:
                    self.errors.append({
                        'batch': batch_num,
                        'error': str(e),
                        'count': len(batch)
                    })
                return {
                    'batch_num': batch_num,
                    'imported': 0,
                    'errors': [{'error': str(e), 'count': len(batch)}]
                }
    
    def import_contacts(self):
        """Main import function with batch processing and threading"""
        self.stats['start_time'] = datetime.now()
        logger.info("=" * 80)
        logger.info("Starting batch contact import")
        logger.info(f"File: {self.excel_path}")
        logger.info(f"Batch size: {self.batch_size}")
        logger.info(f"Threads: {self.num_threads}")
        logger.info(f"Update existing: {self.update_existing}")
        logger.info(f"Import only new: {self.import_only_new}")
        logger.info(f"Dry run: {self.dry_run}")
        logger.info("=" * 80)
        
        try:
            # Load contacts from Excel
            contacts = self.load_excel_file()
            
            if not contacts:
                logger.warning("No contacts found in Excel file")
                return
            
            # Split into batches
            batches = []
            for i in range(0, len(contacts), self.batch_size):
                batch = contacts[i:i + self.batch_size]
                batches.append((batch, len(batches) + 1))
            
            self.stats['batches'] = len(batches)
            logger.info(f"Split into {len(batches)} batches")
            
            # Process batches with threading
            with ThreadPoolExecutor(max_workers=self.num_threads) as executor:
                # Submit all batches
                future_to_batch = {
                    executor.submit(self.process_batch, batch, batch_num): batch_num
                    for batch, batch_num in batches
                }
                
                # Process completed batches
                for future in as_completed(future_to_batch):
                    batch_num = future_to_batch[future]
                    try:
                        result = future.result()
                        logger.debug(f"Batch {batch_num} completed: {result['imported']} imported")
                    except Exception as e:
                        logger.error(f"Batch {batch_num} exception: {e}", exc_info=True)
            
            self.stats['end_time'] = datetime.now()
            duration = (self.stats['end_time'] - self.stats['start_time']).total_seconds()
            
            # Print summary
            logger.info("=" * 80)
            logger.info("IMPORT SUMMARY")
            logger.info("=" * 80)
            logger.info(f"Total rows: {self.stats['total_rows']}")
            logger.info(f"Processed: {self.stats['processed']}")
            logger.info(f"Imported: {self.stats['imported']}")
            logger.info(f"Errors: {self.stats['errors']}")
            logger.info(f"Batches: {self.stats['batches']}")
            logger.info(f"Duration: {duration:.2f} seconds")
            logger.info(f"Rate: {self.stats['processed'] / duration:.2f} contacts/second")
            logger.info("=" * 80)
            
            if self.errors:
                logger.warning(f"Total errors: {len(self.errors)}")
                for error in self.errors[:10]:  # Show first 10 errors
                    logger.warning(f"Error: {error}")
            
        except Exception as e:
            logger.error(f"Import failed: {e}", exc_info=True)
            raise


def main():
    parser = argparse.ArgumentParser(description='Batch import contacts from Excel file')
    parser.add_argument('excel_file', help='Path to Excel file')
    parser.add_argument('--batch-size', type=int, default=100, help='Number of records per batch (default: 100)')
    parser.add_argument('--threads', type=int, default=4, help='Number of parallel threads (default: 4)')
    parser.add_argument('--update-existing', action='store_true', help='Update existing contacts')
    parser.add_argument('--import-only-new', action='store_true', help='Only import new contacts, skip duplicates')
    parser.add_argument('--sheet', type=str, help='Specific sheet name to import (default: all sheets)')
    parser.add_argument('--dry-run', action='store_true', help='Preview mode, don\'t actually import')
    
    args = parser.parse_args()
    
    importer = BatchContactImporter(
        excel_path=args.excel_file,
        batch_size=args.batch_size,
        num_threads=args.threads,
        update_existing=args.update_existing,
        import_only_new=args.import_only_new,
        sheet_name=args.sheet,
        dry_run=args.dry_run
    )
    
    try:
        importer.import_contacts()
        logger.info("Import completed successfully!")
    except Exception as e:
        logger.error(f"Import failed: {e}")
        sys.exit(1)


if __name__ == '__main__':
    main()




