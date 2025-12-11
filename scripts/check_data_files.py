"""Check data files for row counts"""
import os
import sys
from pathlib import Path
import openpyxl
import csv

data_dir = Path(__file__).parent.parent / 'data'

print("\n=== DATA FILES ANALYSIS ===\n")

total_rows = 0

# Check Excel files
xlsx_files = list(data_dir.glob('*.xlsx'))
for file_path in xlsx_files:
    print(f"[EXCEL] {file_path.name}")
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        file_total = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = ws.max_row - 1  # Subtract header
            file_total += rows
            total_rows += rows
            print(f"   Sheet: {sheet_name} - {rows} rows")
        print(f"   TOTAL: {file_total} rows\n")
    except Exception as e:
        print(f"   ERROR: {e}\n")

# Check CSV files
csv_files = list(data_dir.glob('*.csv'))
for file_path in csv_files:
    print(f"[CSV] {file_path.name}")
    try:
        # Try different encodings
        for encoding in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']:
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    rows = sum(1 for _ in f) - 1  # Subtract header
                    total_rows += rows
                    print(f"   Rows: {rows} (encoding: {encoding})\n")
                    break
            except UnicodeDecodeError:
                continue
    except Exception as e:
        print(f"   ERROR: {e}\n")

print(f"{'='*50}")
print(f"GRAND TOTAL: {total_rows} rows expected")
print(f"{'='*50}\n")

