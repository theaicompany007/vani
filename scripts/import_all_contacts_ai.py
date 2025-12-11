"""
AI-Powered Bulk Contact Import
- Scans directory for all .xlsx and .csv files
- Handles multiple sheets per file
- Deduplicates across all files
- Uses AI to infer industry when missing
"""
import os
import sys
import logging
from pathlib import Path
from typing import List, Dict, Any
import glob
from dotenv import load_dotenv
from datetime import datetime

# Add parent directory to path
basedir = Path(__file__).parent.parent
sys.path.insert(0, str(basedir))

# Load environment
load_dotenv(basedir / '.env')
load_dotenv(basedir / '.env.local', override=True)

from app.supabase_client import init_supabase
from app.services.contact_service import upsert_contacts, normalize_email, normalize_phone

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

class AIContactImporter:
    def __init__(self, data_directory='data', batch_size=50, use_ai=True):
        self.data_directory = Path(data_directory)
        self.batch_size = batch_size
        self.use_ai = use_ai
        self.openai_client = None
        
        # Initialize OpenAI if available
        if self.use_ai:
            try:
                import openai
                api_key = os.getenv('OPENAI_API_KEY')
                if api_key:
                    self.openai_client = openai.OpenAI(api_key=api_key)
                    logger.info("OpenAI client initialized for AI industry inference")
                else:
                    logger.warning("OPENAI_API_KEY not found - AI features disabled")
                    self.use_ai = False
            except ImportError:
                logger.warning("OpenAI module not available - AI features disabled")
                self.use_ai = False
        
        self.stats = {
            'files_processed': 0,
            'sheets_processed': 0,
            'total_rows': 0,
            'skipped_empty': 0,
            'duplicates_removed': 0,
            'imported': 0,
            'errors': 0,
            'ai_industries_assigned': 0
        }
    
    def infer_industry_with_ai(self, contact: Dict[str, Any]) -> str:
        """Use AI to infer industry from contact data"""
        if not self.use_ai or not self.openai_client:
            return None
        
        try:
            # Build context from contact data
            context_parts = []
            if contact.get('company'):
                context_parts.append(f"Company: {contact['company']}")
            if contact.get('role'):
                context_parts.append(f"Role: {contact['role']}")
            if contact.get('domain'):
                context_parts.append(f"Domain: {contact['domain']}")
            if contact.get('linkedin'):
                context_parts.append(f"LinkedIn: {contact['linkedin']}")
            
            if not context_parts:
                return None
            
            context = "\n".join(context_parts)
            
            prompt = f"""Based on the following contact information, infer the most likely industry sector. 
Return ONLY ONE lowercase industry name (e.g., 'technology', 'healthcare', 'education', 'finance', 'retail', 'manufacturing', 'consulting', 'real estate', 'hospitality', 'media', 'energy', 'agriculture', 'automotive', 'pharmaceutical', 'telecom', 'logistics', 'construction', 'legal', 'government', 'nonprofit', 'entertainment', 'sports', 'fashion', 'food & beverages', 'fmcg', 'edtech', 'fintech', 'saas', 'ecommerce', 'cryptocurrency', 'aerospace', 'defense', 'chemicals', 'insurance', 'banking', 'investment', 'other').

Contact Information:
{context}

Industry (one word, lowercase):"""

            response = self.openai_client.chat.completions.create(
                model="gpt-3.5-turbo",
                messages=[
                    {"role": "system", "content": "You are an expert at classifying companies into industry sectors. Always respond with just one lowercase industry name."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.3,
                max_tokens=20
            )
            
            industry = response.choices[0].message.content.strip().lower()
            logger.debug(f"AI inferred industry '{industry}' for {contact.get('company', 'unknown')}")
            return industry
            
        except Exception as e:
            logger.warning(f"AI industry inference failed: {e}")
            return None
    
    def load_excel_file(self, file_path: Path) -> List[Dict[str, Any]]:
        """Load all sheets from an Excel file"""
        try:
            from openpyxl import load_workbook
            
            workbook = load_workbook(filename=str(file_path), read_only=True, data_only=True)
            all_contacts = []
            
            for sheet_name in workbook.sheetnames:
                logger.info(f"  Processing sheet: {sheet_name}")
                sheet = workbook[sheet_name]
                
                # Get headers from first row
                headers = []
                for cell in sheet[1]:
                    header = str(cell.value).strip().lower() if cell.value else None
                    if header:
                        # Normalize header names
                        if header in ['name', 'full_name', 'contact_name', 'person']:
                            header = 'name'
                        elif header in ['email', 'email_address', 'e_mail', 'mail']:
                            header = 'email'
                        elif header in ['phone', 'phone_number', 'mobile', 'tel', 'telephone']:
                            header = 'phone'
                        elif header in ['company', 'company_name', 'organization', 'org']:
                            header = 'company'
                        elif header in ['role', 'title', 'job_title', 'position', 'designation']:
                            header = 'role'
                        elif header in ['linkedin', 'linkedin_url', 'linkedin_profile', 'linkedin profile']:
                            header = 'linkedin'
                        elif header in ['industry', 'sector', 'vertical']:
                            header = 'industry'
                        elif header in ['city', 'location', 'place']:
                            header = 'city'
                        elif header in ['lead_source', 'source', 'source_name', 'lead source']:
                            header = 'lead_source'
                    headers.append(header)
                
                # Parse rows
                row_count = 0
                for row in sheet.iter_rows(min_row=2, values_only=False):
                    row_count += 1
                    contact_data = {}
                    for idx, cell in enumerate(row):
                        if idx < len(headers) and headers[idx] and cell.value:
                            contact_data[headers[idx]] = str(cell.value).strip()
                    
                    # Skip completely empty rows (must have at least email or phone to be valid)
                    has_email = bool(contact_data.get('email') and str(contact_data['email']).strip())
                    has_phone = bool(contact_data.get('phone') and str(contact_data['phone']).strip())
                    
                    if not has_email and not has_phone:
                        self.stats['skipped_empty'] += 1
                        continue
                    
                    # Add metadata
                    contact_data['_source_file'] = file_path.name
                    contact_data['_source_sheet'] = sheet_name
                    contact_data['lead_source'] = contact_data.get('lead_source') or sheet_name
                    
                    all_contacts.append(contact_data)
                
                logger.info(f"    Extracted {len([c for c in all_contacts if c.get('_source_sheet') == sheet_name])} contacts from sheet '{sheet_name}'")
                self.stats['sheets_processed'] += 1
            
            workbook.close()
            return all_contacts
            
        except Exception as e:
            logger.error(f"Error loading Excel file {file_path}: {e}")
            return []
    
    def load_csv_file(self, file_path: Path) -> List[Dict[str, Any]]:
        """Load contacts from a CSV file"""
        try:
            import csv
            
            all_contacts = []
            # Try multiple encodings
            encodings = ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
            csvfile = None
            
            for encoding in encodings:
                try:
                    csvfile = open(file_path, 'r', encoding=encoding)
                    # Try to read first line to verify encoding works
                    csvfile.readline()
                    csvfile.seek(0)
                    logger.debug(f"Successfully opened with encoding: {encoding}")
                    break
                except (UnicodeDecodeError, Exception):
                    if csvfile:
                        csvfile.close()
                    continue
            
            if not csvfile:
                logger.error(f"Could not open {file_path} with any supported encoding")
                return []
            
            try:
                # Try to detect dialect
                sample = csvfile.read(8192)
                csvfile.seek(0)
                dialect = csv.Sniffer().sniff(sample)
            except Exception:
                # Use default dialect
                dialect = csv.excel
                logger.debug("Using default CSV dialect")
            
            reader = csv.DictReader(csvfile, dialect=dialect)
            
            for row in reader:
                    # Normalize keys
                    contact_data = {}
                    for key, value in row.items():
                        if not key or not value:
                            continue
                        
                        key_lower = key.strip().lower()
                        value = str(value).strip()
                        
                        # Normalize field names
                        if key_lower in ['name', 'full_name', 'contact_name', 'person']:
                            contact_data['name'] = value
                        elif key_lower in ['email', 'email_address', 'e_mail', 'mail']:
                            contact_data['email'] = value
                        elif key_lower in ['phone', 'phone_number', 'mobile', 'tel']:
                            contact_data['phone'] = value
                        elif key_lower in ['company', 'company_name', 'organization']:
                            contact_data['company'] = value
                        elif key_lower in ['role', 'title', 'job_title', 'position']:
                            contact_data['role'] = value
                        elif key_lower in ['linkedin', 'linkedin_url', 'linkedin_profile']:
                            contact_data['linkedin'] = value
                        elif key_lower in ['industry', 'sector']:
                            contact_data['industry'] = value
                        elif key_lower in ['city', 'location']:
                            contact_data['city'] = value
                        elif key_lower in ['lead_source', 'source', 'source_name']:
                            contact_data['lead_source'] = value
                        else:
                            contact_data[key_lower] = value
                    
                    # Skip completely empty rows (must have at least email or phone to be valid)
                    has_email = bool(contact_data.get('email') and str(contact_data['email']).strip())
                    has_phone = bool(contact_data.get('phone') and str(contact_data['phone']).strip())
                    
                    if not has_email and not has_phone:
                        self.stats['skipped_empty'] += 1
                        continue
                    
                    contact_data['_source_file'] = file_path.name
                    contact_data['lead_source'] = contact_data.get('lead_source') or file_path.stem
                    all_contacts.append(contact_data)
            
            csvfile.close()
            logger.info(f"  Extracted {len(all_contacts)} contacts from CSV")
            self.stats['sheets_processed'] += 1
            return all_contacts
            
        except Exception as e:
            logger.error(f"Error loading CSV file {file_path}: {e}")
            return []
    
    def deduplicate_contacts(self, contacts: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Merge duplicates based on email/phone, keeping additional contact info
        If same person appears in multiple sheets, merge email2, mobile2, etc.
        """
        contact_map = {}  # key: (email, phone) -> contact data
        duplicates = 0
        
        for contact in contacts:
            email = normalize_email(contact.get('email'))
            phone = normalize_phone(contact.get('phone'))
            
            # Create a unique key (use email as primary, phone as fallback)
            key = email if email else phone
            
            if not key:
                continue  # Skip if no email or phone
            
            if key in contact_map:
                # Duplicate found - merge additional info
                duplicates += 1
                existing = contact_map[key]
                
                # Merge phone numbers (if different)
                if phone and phone != normalize_phone(existing.get('phone')):
                    if not existing.get('phone2'):
                        existing['phone2'] = phone
                    elif not existing.get('phone3'):
                        existing['phone3'] = phone
                
                # Merge emails (if different)
                existing_email = normalize_email(existing.get('email'))
                if email and email != existing_email:
                    if not existing.get('email2'):
                        existing['email2'] = email
                    elif not existing.get('email3'):
                        existing['email3'] = email
                
                # Keep more complete data (prefer contact with more fields)
                if not existing.get('name') and contact.get('name'):
                    existing['name'] = contact['name']
                if not existing.get('company') and contact.get('company'):
                    existing['company'] = contact['company']
                if not existing.get('role') and contact.get('role'):
                    existing['role'] = contact['role']
                if not existing.get('industry') and contact.get('industry'):
                    existing['industry'] = contact['industry']
                if not existing.get('city') and contact.get('city'):
                    existing['city'] = contact['city']
                if not existing.get('linkedin') and contact.get('linkedin'):
                    existing['linkedin'] = contact['linkedin']
                
                logger.debug(f"Merged duplicate: {contact.get('name')} ({key})")
            else:
                # New contact
                contact_map[key] = contact
        
        unique_contacts = list(contact_map.values())
        self.stats['duplicates_removed'] = duplicates
        logger.info(f"Merged {duplicates} duplicates, {len(unique_contacts)} unique contacts remain")
        return unique_contacts
    
    def process_all_files(self):
        """Find and process all Excel and CSV files"""
        logger.info(f"\n{'='*70}")
        logger.info(f"AI-POWERED BULK CONTACT IMPORT")
        logger.info(f"{'='*70}\n")
        logger.info(f"Scanning directory: {self.data_directory}")
        
        # Find all files
        excel_files = list(self.data_directory.glob('*.xlsx')) + list(self.data_directory.glob('*.xls'))
        csv_files = list(self.data_directory.glob('*.csv'))
        all_files = excel_files + csv_files
        
        if not all_files:
            logger.error(f"No .xlsx, .xls, or .csv files found in {self.data_directory}")
            return False
        
        logger.info(f"Found {len(all_files)} file(s): {len(excel_files)} Excel, {len(csv_files)} CSV\n")
        
        # Load all contacts from all files
        all_contacts = []
        for file_path in all_files:
            logger.info(f"Processing: {file_path.name}")
            
            if file_path.suffix.lower() in ['.xlsx', '.xls']:
                contacts = self.load_excel_file(file_path)
            else:
                contacts = self.load_csv_file(file_path)
            
            all_contacts.extend(contacts)
            self.stats['files_processed'] += 1
        
        logger.info(f"\nTotal contacts extracted: {len(all_contacts)}")
        self.stats['total_rows'] = len(all_contacts)
        
        # Deduplicate
        logger.info("\nDeduplicating contacts...")
        unique_contacts = self.deduplicate_contacts(all_contacts)
        
        # Process industries with AI
        logger.info("\nProcessing industries...")
        for contact in unique_contacts:
            if not contact.get('industry'):
                inferred_industry = self.infer_industry_with_ai(contact)
                if inferred_industry:
                    contact['industry'] = inferred_industry
                    self.stats['ai_industries_assigned'] += 1
            else:
                # Normalize existing industry to lowercase
                contact['industry'] = contact['industry'].lower()
        
        logger.info(f"AI assigned {self.stats['ai_industries_assigned']} industries")
        
        # Import to database
        logger.info(f"\nImporting {len(unique_contacts)} contacts to database...")
        supabase = init_supabase(os.getenv('SUPABASE_URL'), os.getenv('SUPABASE_SERVICE_KEY'))
        
        if not supabase:
            logger.error("Failed to initialize Supabase client")
            return False
        
        # Process in batches
        for i in range(0, len(unique_contacts), self.batch_size):
            batch = unique_contacts[i:i + self.batch_size]
            batch_num = (i // self.batch_size) + 1
            total_batches = (len(unique_contacts) + self.batch_size - 1) // self.batch_size
            
            logger.info(f"Importing batch {batch_num}/{total_batches}: {len(batch)} contacts")
            
            try:
                result = upsert_contacts(supabase, batch, {'updateExisting': False})
                self.stats['imported'] += result.get('imported', 0)
                self.stats['errors'] += len(result.get('errors', []))
            except Exception as e:
                logger.error(f"Batch {batch_num} error: {e}")
                self.stats['errors'] += len(batch)
        
        # Print summary
        logger.info(f"\n{'='*70}")
        logger.info(f"IMPORT SUMMARY")
        logger.info(f"{'='*70}")
        logger.info(f"Files processed: {self.stats['files_processed']}")
        logger.info(f"Sheets processed: {self.stats['sheets_processed']}")
        logger.info(f"Total rows: {self.stats['total_rows']}")
        logger.info(f"Duplicates removed: {self.stats['duplicates_removed']}")
        logger.info(f"AI industries assigned: {self.stats['ai_industries_assigned']}")
        logger.info(f"Successfully imported: {self.stats['imported']}")
        logger.info(f"Errors: {self.stats['errors']}")
        logger.info(f"{'='*70}\n")
        
        return True

def main():
    import argparse
    
    parser = argparse.ArgumentParser(description='AI-powered bulk contact import')
    parser.add_argument('--directory', '-d', default='data', help='Directory containing Excel/CSV files (default: data)')
    parser.add_argument('--batch-size', '-b', type=int, default=50, help='Batch size for import (default: 50)')
    parser.add_argument('--no-ai', action='store_true', help='Disable AI industry inference')
    parser.add_argument('--clear-first', action='store_true', help='Clear existing contacts and companies before import')
    
    args = parser.parse_args()
    
    # Clear tables if requested
    if args.clear_first:
        logger.info("Clearing existing contacts and companies...")
        supabase = init_supabase(os.getenv('SUPABASE_URL'), os.getenv('SUPABASE_SERVICE_KEY'))
        if supabase:
            try:
                supabase.table('contacts').delete().neq('id', '00000000-0000-0000-0000-000000000000').execute()
                logger.info("  Cleared contacts")
                supabase.table('companies').delete().neq('id', '00000000-0000-0000-0000-000000000000').execute()
                logger.info("  Cleared companies")
            except Exception as e:
                logger.error(f"Error clearing tables: {e}")
    
    # Run import
    importer = AIContactImporter(
        data_directory=args.directory,
        batch_size=args.batch_size,
        use_ai=not args.no_ai
    )
    
    success = importer.process_all_files()
    sys.exit(0 if success else 1)

if __name__ == '__main__':
    main()

