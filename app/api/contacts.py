"""Contacts API routes"""
from flask import Blueprint, request, jsonify, current_app
import logging
import openpyxl
from io import BytesIO
from app.models.contacts import Contact
from app.supabase_client import get_supabase_client
from app.auth import require_auth, require_use_case, get_current_industry, get_supabase_auth_client
from app.services.contact_service import (
    upsert_contacts, find_duplicates, resolve_best_domain,
    normalize_email, normalize_phone
)
from uuid import UUID

logger = logging.getLogger(__name__)

contacts_bp = Blueprint('contacts', __name__)


@contacts_bp.route('/api/contacts', methods=['GET'])
@require_auth
@require_use_case('contact_management')
def list_contacts():
    """List all contacts with optional filters and industry-based access control"""
    try:
        supabase = get_supabase_client(current_app)
        if not supabase:
            return jsonify({'error': 'Supabase not configured'}), 503
        
        # Get current user for industry-based filtering
        from app.auth import get_current_user
        user = get_current_user()
        user_industry_id = None
        is_industry_admin = False
        is_super_user = False
        
        if user:
            is_industry_admin = getattr(user, 'is_industry_admin', False)
            is_super_user = getattr(user, 'is_super_user', False)
            if hasattr(user, 'industry_id') and user.industry_id:
                user_industry_id = str(user.industry_id)
            elif hasattr(user, 'active_industry_id') and user.active_industry_id:
                user_industry_id = str(user.active_industry_id)
        
        # Get query parameters
        company_id = request.args.get('company_id')
        industry = request.args.get('industry')  # Manual filter (super users only)
        search = request.args.get('search')
        limit = int(request.args.get('limit', 1000))  # Increased default limit to show all contacts
        offset = int(request.args.get('offset', 0))
        
        # Industry-based filtering - only apply if user is NOT a super user
        # Super users should see all contacts regardless of industry
        if is_industry_admin and user_industry_id and not is_super_user:
            # Industry admin: Only see contacts from their assigned industry
            # Get industry name from industry_id
            industry_lookup = supabase.table('industries').select('name').eq('id', user_industry_id).limit(1).execute()
            if industry_lookup.data:
                industry = industry_lookup.data[0]['name']  # Override with user's industry
        
        # Get total count first (before filtering)
        # Use left join to include contacts without companies
        count_query = supabase.table('contacts').select('id', count='exact')
        if company_id:
            count_query = count_query.eq('company_id', company_id)
        if industry:
            count_query = count_query.ilike('industry', f'%{industry}%')
        if search:
            count_query = count_query.or_(f'name.ilike.%{search}%,email.ilike.%{search}%,phone.ilike.%{search}%,lead_source.ilike.%{search}%,role.ilike.%{search}%,company.ilike.%{search}%,industry.ilike.%{search}%')
        # Execute count query - Supabase returns count in response
        count_response = count_query.execute()
        # Try multiple ways to get the count
        if hasattr(count_response, 'count') and count_response.count is not None:
            total_count = count_response.count
        elif hasattr(count_response, 'data') and count_response.data:
            total_count = len(count_response.data)
        else:
            # Fallback: query all matching records to count
            all_query = supabase.table('contacts').select('id')
            if company_id:
                all_query = all_query.eq('company_id', company_id)
            if industry:
                all_query = all_query.ilike('industry', f'%{industry}%')
            if search:
                all_query = all_query.or_(f'name.ilike.%{search}%,email.ilike.%{search}%,phone.ilike.%{search}%,lead_source.ilike.%{search}%,role.ilike.%{search}%,company.ilike.%{search}%,industry.ilike.%{search}%')
            all_response = all_query.execute()
            total_count = len(all_response.data) if all_response.data else 0
        
        # Select contacts with company information (using left join to include contacts without companies)
        # Use left join syntax: companies!left(...) to include contacts without company_id
        query = supabase.table('contacts').select('*, companies!left(name, industry, domain)')
        
        if company_id:
            query = query.eq('company_id', company_id)
        if industry:
            query = query.ilike('industry', f'%{industry}%')
        if search:
            query = query.or_(f'name.ilike.%{search}%,email.ilike.%{search}%,phone.ilike.%{search}%,lead_source.ilike.%{search}%,role.ilike.%{search}%,company.ilike.%{search}%,industry.ilike.%{search}%')
        
        query = query.order('created_at', desc=True).limit(limit).offset(offset)
        response = query.execute()
        
        contacts = []
        for c in response.data:
            try:
                contact = Contact.from_dict(c).to_dict()
                # Enrich with company name if available
                # Handle both single company object and array of companies
                companies_data = c.get('companies')
                if companies_data:
                    if isinstance(companies_data, list) and len(companies_data) > 0:
                        # Array format (from join)
                        company = companies_data[0]
                        if company and company.get('name'):
                            contact['company_name'] = company['name']
                        if company and company.get('industry'):
                            contact['company_industry'] = company['industry']
                    elif isinstance(companies_data, dict):
                        # Single object format
                        if companies_data.get('name'):
                            contact['company_name'] = companies_data['name']
                        if companies_data.get('industry'):
                            contact['company_industry'] = companies_data['industry']
                contacts.append(contact)
            except Exception as contact_error:
                logger.warning(f"Error processing contact {c.get('id', 'unknown')}: {contact_error}")
                import traceback
                logger.warning(traceback.format_exc())
                # Skip invalid contacts but continue processing
                continue
        
        # Log summary for debugging
        contacts_with_email = len([c for c in contacts if c.get('email')])
        logger.info(f"Contacts API: Returning {len(contacts)} contacts (total: {total_count}), {contacts_with_email} with emails")
        
        return jsonify({
            'success': True,
            'contacts': contacts,
            'count': len(contacts),
            'total': total_count
        })
        
    except Exception as e:
        logger.error(f"Error listing contacts: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500


@contacts_bp.route('/api/contacts/<contact_id>', methods=['GET'])
@require_auth
@require_use_case('contact_management')
def get_contact(contact_id):
    """Get a single contact"""
    try:
        # Validate UUID format
        try:
            contact_uuid = UUID(contact_id)
        except (ValueError, TypeError):
            return jsonify({'error': 'Invalid contact ID format'}), 400
        
        supabase = get_supabase_client(current_app)
        if not supabase:
            return jsonify({'error': 'Supabase not configured'}), 503
        
        response = supabase.table('contacts').select('*').eq('id', str(contact_uuid)).limit(1).execute()
        
        if not response.data:
            return jsonify({'error': 'Contact not found'}), 404
        
        contact = Contact.from_dict(response.data[0])
        
        return jsonify({
            'success': True,
            'contact': contact.to_dict()
        })
        
    except Exception as e:
        logger.error(f"Error getting contact {contact_id}: {e}")
        return jsonify({'error': str(e)}), 500


@contacts_bp.route('/api/contacts', methods=['POST'])
@require_auth
@require_use_case('contact_management')
def create_contact():
    """Create or update a contact (upsert by email)"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        supabase = get_supabase_client(current_app)
        if not supabase:
            return jsonify({'error': 'Supabase not configured'}), 503
        
        # Resolve company_id if company name/domain provided
        domain = resolve_best_domain(data.get('domain'), data.get('email'))
        company_name = data.get('company') or data.get('company_name')
        industry = data.get('industry')
        
        if domain or company_name:
            company_id = find_or_create_company(supabase, company_name, domain, industry)
            if company_id:
                data['company_id'] = company_id
        
        # Normalize email and phone
        if data.get('email'):
            data['email'] = normalize_email(data['email'])
        if data.get('phone'):
            data['phone'] = normalize_phone(data['phone'])
        
        # Normalize lead_source
        if data.get('leadSource') and not data.get('lead_source'):
            data['lead_source'] = data['leadSource']
        if not data.get('lead_source') and not data.get('leadSource') and data.get('sheet'):
            data['lead_source'] = data['sheet']
        
        # Upsert by email if email provided
        if data.get('email'):
            email_norm = normalize_email(data['email'])
            # Check if exists
            existing = supabase.table('contacts').select('*').eq('email', email_norm).limit(1).execute()
            
            if existing.data and len(existing.data) > 0:
                # Update existing
                contact_id = existing.data[0]['id']
                response = supabase.table('contacts').update(data).eq('id', contact_id).execute()
            else:
                # Insert new
                response = supabase.table('contacts').insert(data).execute()
        else:
            # Insert new if no email
            response = supabase.table('contacts').insert(data).execute()
        
        if not response.data:
            return jsonify({'error': 'Failed to create/update contact'}), 500
        
        contact = Contact.from_dict(response.data[0])
        
        return jsonify({
            'success': True,
            'contact': contact.to_dict()
        }), 201
        
    except Exception as e:
        logger.error(f"Error creating/updating contact: {e}")
        return jsonify({'error': str(e)}), 500


@contacts_bp.route('/api/contacts/<contact_id>', methods=['PUT', 'PATCH'])
@require_auth
@require_use_case('contact_management')
def update_contact(contact_id):
    """Update a contact"""
    try:
        # Validate UUID format
        try:
            contact_uuid = UUID(contact_id)
        except (ValueError, TypeError):
            return jsonify({'error': 'Invalid contact ID format'}), 400
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        supabase = get_supabase_client(current_app)
        if not supabase:
            return jsonify({'error': 'Supabase not configured'}), 503
        
        # Remove id from update data
        update_data = {k: v for k, v in data.items() if k != 'id'}
        
        # Normalize email and phone if provided
        if 'email' in update_data and update_data['email']:
            update_data['email'] = normalize_email(update_data['email'])
        if 'phone' in update_data and update_data['phone']:
            update_data['phone'] = normalize_phone(update_data['phone'])
        
        response = supabase.table('contacts').update(update_data).eq('id', str(contact_uuid)).execute()
        
        if not response.data:
            return jsonify({'error': 'Contact not found'}), 404
        
        updated_contact = Contact.from_dict(response.data[0])
        
        return jsonify({
            'success': True,
            'contact': updated_contact.to_dict()
        })
        
    except Exception as e:
        logger.error(f"Error updating contact {contact_id}: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@contacts_bp.route('/api/contacts/<contact_id>', methods=['DELETE'])
@require_auth
@require_use_case('contact_management')
def delete_contact(contact_id):
    """Delete a contact"""
    try:
        # Validate UUID format
        try:
            contact_uuid = UUID(contact_id)
        except (ValueError, TypeError):
            return jsonify({'error': 'Invalid contact ID format'}), 400
        
        supabase = get_supabase_client(current_app)
        if not supabase:
            return jsonify({'error': 'Supabase not configured'}), 503
        
        response = supabase.table('contacts').delete().eq('id', str(contact_uuid)).execute()
        
        return jsonify({
            'success': True,
            'message': 'Contact deleted successfully'
        })
        
    except Exception as e:
        logger.error(f"Error deleting contact {contact_id}: {e}")
        return jsonify({'error': str(e)}), 500


@contacts_bp.route('/api/contacts/bulk', methods=['POST'])
@require_auth
@require_use_case('contact_management')
def bulk_import_contacts():
    """Bulk import contacts from JSON array"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        supabase = get_supabase_client(current_app)
        if not supabase:
            return jsonify({'error': 'Supabase not configured'}), 503
        
        # Get contacts array
        contacts = data.get('contacts', [])
        if not isinstance(contacts, list):
            contacts = [contacts] if contacts else []
        
        if not contacts:
            return jsonify({'error': 'No contacts provided'}), 400
        
        # Get options
        is_preview = data.get('preview', False)
        is_commit = data.get('commit', False)
        update_existing = data.get('updateExisting', False)
        force_commit = data.get('force', False)
        provided_sheet = data.get('sheet')
        
        # Normalize rows
        normalized_rows = []
        for row in contacts:
            normalized = {**row}
            
            # Extract domain from email if not provided
            if not normalized.get('domain') and normalized.get('email'):
                normalized['domain'] = resolve_best_domain(None, normalized['email'])
            
            # Set sheet name
            if provided_sheet:
                normalized['sheet'] = provided_sheet
            elif not normalized.get('sheet'):
                normalized['sheet'] = 'import'
            
            # Normalize lead_source
            if normalized.get('leadSource') and not normalized.get('lead_source'):
                normalized['lead_source'] = normalized['leadSource']
            if not normalized.get('lead_source') and not normalized.get('leadSource') and normalized.get('sheet'):
                normalized['lead_source'] = normalized['sheet']
            
            normalized_rows.append(normalized)
        
        # Preview mode: check for duplicates
        if is_preview:
            duplicates, uniques = find_duplicates(supabase, normalized_rows)
            
            # Sample of transformed rows
            sample = normalized_rows[:5]
            
            breakdown = {
                'emailMatches': len([d for d in duplicates if d.get('match_type') == 'email']),
                'phoneMatches': len([d for d in duplicates if d.get('match_type') == 'phone'])
            }
            
            if not is_commit:
                return jsonify({
                    'preview': True,
                    'counts': {
                        'total': len(normalized_rows),
                        'duplicates': len(duplicates),
                        'uniques': len(uniques),
                        'toBeUpdated': len(duplicates),
                        'toBeInserted': len(uniques)
                    },
                    'breakdown': breakdown,
                    'duplicates': duplicates[:10],  # Limit for response size
                    'uniques': uniques[:10],
                    'sample': sample
                })
        
        # Commit mode
        if is_commit:
            # Enforce preview step unless force=true
            if not is_preview and not force_commit:
                return jsonify({
                    'error': 'Commit requires a prior preview. Send preview=true to inspect duplicates/uniques, then commit=true to import. To override, send force=true.'
                }), 400
            
            from datetime import datetime
            import_start_time = datetime.now()
            logger.info(f"Starting import of {len(normalized_rows)} contacts...")
            result = upsert_contacts(supabase, normalized_rows, {'updateExisting': update_existing})
            import_duration = (datetime.now() - import_start_time).total_seconds()
            logger.info(f"Import completed in {import_duration:.2f}s: {result.get('imported', 0)} imported, {len(result.get('errors', []))} errors")
            
            if result.get('errors') and len(result['errors']) > 0:
                return jsonify({
                    'imported': result.get('imported', 0),
                    'errors': result.get('errors', []),
                    'data': result.get('data', []),
                    'report': result.get('report', [])
                }), 400
            
            return jsonify({
                'imported': result.get('imported', 0),
                'data': result.get('data', []),
                'report': result.get('report', [])
            })
        
        return jsonify({
            'error': 'Commit flag not provided. Send preview=true to inspect duplicates, then commit=true to import.'
        }), 400
        
    except Exception as e:
        logger.error(f"Error in bulk import: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@contacts_bp.route('/api/contacts/import-excel', methods=['POST'])
@require_auth
@require_use_case('contact_management')
def import_excel_contacts():
    """Import contacts from Excel file (multi-sheet support) - supports background jobs for large files"""
    try:
        # Enhanced error logging
        logger.info(f"Import request received - Content-Type: {request.content_type}, Content-Length: {request.content_length}")
        
        if 'file' not in request.files:
            logger.warning("Import request missing 'file' in request.files")
            return jsonify({'error': 'No file provided', 'details': 'The file field is missing from the request'}), 400
        
        file = request.files['file']
        if file.filename == '':
            logger.warning("Import request with empty filename")
            return jsonify({'error': 'No file selected', 'details': 'The file name is empty'}), 400
        
        # Check file size
        file.seek(0, 2)  # Seek to end
        file_size = file.tell()
        file.seek(0)  # Reset to beginning
        
        logger.info(f"File upload: {file.filename}, Size: {file_size} bytes ({file_size / 1024 / 1024:.2f} MB)")
        
        if file_size > 50 * 1024 * 1024:  # 50MB limit
            logger.warning(f"File too large: {file_size} bytes")
            return jsonify({'error': 'File too large', 'details': f'File size ({file_size / 1024 / 1024:.2f} MB) exceeds maximum allowed size (50 MB)'}), 400
        
        # Check file extension
        if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
            logger.warning(f"Invalid file type: {file.filename}")
            return jsonify({'error': 'Invalid file type. Only Excel files (.xlsx, .xls) are supported.'}), 400
        
        supabase = get_supabase_client(current_app)
        if not supabase:
            return jsonify({'error': 'Supabase not configured'}), 503
        
        # Get app_users.id from current user context (not Supabase Auth user_id)
        from app.auth import get_current_user
        from flask import session
        
        current_user = get_current_user()
        app_user_id = None
        
        if current_user:
            # Try to get app_users.id from current user context
            app_user_id = str(getattr(current_user, 'id', None))
            logger.debug(f"Got app_user_id from current_user: {app_user_id}")
        
        # Fallback: look up app_users.id from supabase_user_id in session
        if not app_user_id:
            supabase_user_id = session.get('user_id')
            logger.debug(f"Looking up app_user_id from supabase_user_id: {supabase_user_id}")
            if supabase_user_id:
                try:
                    app_user_response = supabase.table('app_users').select('id').eq('supabase_user_id', supabase_user_id).limit(1).execute()
                    if app_user_response.data and len(app_user_response.data) > 0:
                        app_user_id = str(app_user_response.data[0]['id'])
                        logger.info(f"Resolved app_user_id {app_user_id} from supabase_user_id {supabase_user_id}")
                    else:
                        logger.error(f"app_user not found for supabase_user_id: {supabase_user_id}")
                        # Try to create app_user if it doesn't exist
                        try:
                            # Get user email from Supabase Auth
                            supabase_auth = get_supabase_auth_client()
                            if supabase_auth:
                                auth_user = supabase_auth.auth.get_user(session.get('access_token'))
                                if auth_user and auth_user.user:
                                    email = auth_user.user.email
                                    # Create app_user
                                    new_app_user = supabase.table('app_users').insert({
                                        'supabase_user_id': supabase_user_id,
                                        'email': email,
                                        'name': email.split('@')[0] if email else 'User',
                                        'is_super_user': False,
                                        'is_industry_admin': False
                                    }).select('id').execute()
                                    if new_app_user.data:
                                        app_user_id = str(new_app_user.data[0]['id'])
                                        logger.info(f"Created new app_user with id: {app_user_id}")
                        except Exception as create_error:
                            logger.error(f"Error creating app_user: {create_error}")
                        
                        if not app_user_id:
                            return jsonify({
                                'error': 'User record not found in app_users table',
                                'hint': 'Please ensure you are logged in and your user exists in the app_users table'
                            }), 404
                except Exception as e:
                    logger.error(f"Error looking up app_user: {e}")
                    import traceback
                    logger.error(traceback.format_exc())
                    return jsonify({'error': 'Failed to resolve user ID'}), 500
            else:
                return jsonify({'error': 'User not authenticated'}), 401
        
        if not app_user_id:
            return jsonify({'error': 'User not authenticated'}), 401
        
        # Use app_user_id (not supabase_user_id)
        user_id = app_user_id
        logger.info(f"Using app_user_id for job creation: {user_id} (supabase_user_id was: {session.get('user_id')})")
        
        # Verify the user_id exists in app_users table before proceeding (only if we have user_id)
        if user_id:
            try:
                verify_user = supabase.table('app_users').select('id').eq('id', user_id).limit(1).execute()
                if not verify_user.data or len(verify_user.data) == 0:
                    logger.error(f"app_user_id {user_id} does not exist in app_users table!")
                    # Clear user_id so we can use supabase_user_id fallback
                    user_id = None
                    logger.warning(f"Will use supabase_user_id fallback instead")
            except Exception as verify_error:
                logger.error(f"Error verifying user_id: {verify_error}")
                # Clear user_id so we can use supabase_user_id fallback
                user_id = None
                logger.warning(f"Will use supabase_user_id fallback instead")
        
        # Get options
        get_headers_only = request.form.get('get_headers', 'false').lower() == 'true'
        is_preview = request.form.get('preview', 'false').lower() == 'true'
        is_commit = request.form.get('commit', 'false').lower() == 'true'
        use_background_job = request.form.get('useBackgroundJob', 'true').lower() == 'true'
        
        # For header-only requests, process synchronously (fast)
        if get_headers_only:
            try:
                file_content = file.read()
                workbook = openpyxl.load_workbook(BytesIO(file_content), data_only=True, read_only=True)
                sheet = workbook[workbook.sheetnames[0]]
                headers = []
                for cell in sheet[1]:
                    header = str(cell.value).strip() if cell.value else ''
                    if header:
                        headers.append(header)
                workbook.close()
                return jsonify({
                    'success': True,
                    'headers': headers
                })
            except Exception as e:
                logger.error(f"Error reading headers: {e}", exc_info=True)
                return jsonify({'error': f'Failed to read Excel file: {str(e)}'}), 400
        
        # For large files or when explicitly requested, use background job
        # Threshold: > 500 contacts or file size > 1MB
        file_content = file.read()
        estimated_contacts = file_size // 500  # Rough estimate: ~500 bytes per contact
        
        # Determine if we should use background job
        # For large files, ALWAYS use background job for commit operations
        should_use_background = (
            use_background_job and 
            (estimated_contacts > 500 or file_size > 1024 * 1024 or is_commit)
        )
        
        # For commit operations with large files, force background job
        if is_commit and (estimated_contacts > 500 or file_size > 1024 * 1024):
            should_use_background = True
            use_background_job = True
        
        if should_use_background and is_commit:
            # Create background job
            import uuid
            from datetime import datetime
            from app.models.import_job import ImportJob
            
            job_id = str(uuid.uuid4())
            
            # Get import options
            column_map = request.form.get('columnMap')
            if column_map:
                import json
                try:
                    column_map = json.loads(column_map)
                except:
                    column_map = {}
            else:
                column_map = {}
            
            options = {
                'updateExisting': request.form.get('updateExisting', 'false').lower() == 'true',
                'importOnlyNew': request.form.get('importOnlyNew', 'false').lower() == 'true',
                'selectedSheets': request.form.get('selectedSheets', ''),
                'columnMap': column_map
            }
            
            # Create job record
            job_data = {
                'id': job_id,
                'status': 'pending',
                'file_name': file.filename,
                'file_size': file_size,
                'options': options,
                'total_records': 0,
                'processed_records': 0,
                'imported_count': 0,
                'error_count': 0,
                'skipped_count': 0
            }
            
            # Only set user_id if we have a valid app_users.id, otherwise use supabase_user_id as fallback
            if user_id:
                job_data['user_id'] = user_id
            else:
                # Store supabase_user_id as fallback if app_user_id not available
                supabase_user_id_fallback = session.get('user_id')
                if supabase_user_id_fallback:
                    job_data['supabase_user_id'] = supabase_user_id_fallback
                    logger.warning(f"Using supabase_user_id as fallback for job {job_id} (app_user_id not available)")
            
            try:
                insert_response = supabase.table('import_jobs').insert(job_data).execute()
                if not insert_response.data:
                    logger.error(f"Job record insert returned no data for job {job_id}")
                    # Try to verify it was created
                    verify_response = supabase.table('import_jobs').select('id').eq('id', job_id).limit(1).execute()
                    if not verify_response.data:
                        logger.error(f"Job {job_id} was not created in database")
                        # Return error instead of continuing
                        return jsonify({
                            'error': 'Failed to create import job. Please ensure the import_jobs table exists.',
                            'hint': 'Run the migration: app/migrations/010_import_jobs_table.sql in Supabase SQL Editor'
                        }), 500
            except Exception as e:
                logger.error(f"Error creating job record: {e}")
                import traceback
                logger.error(traceback.format_exc())
                # Return error instead of continuing with invalid job_id
                return jsonify({
                    'error': f'Failed to create import job: {str(e)}',
                    'hint': 'Please ensure the import_jobs table exists. Run migration: app/migrations/010_import_jobs_table.sql'
                }), 500
            
            # Start background job
            from app.jobs.import_job import process_import_job
            process_import_job(job_id, file_content, file.filename, column_map, options, current_app.app_context())
            
            logger.info(f"Started background import job {job_id} for file {file.filename}")
            
            return jsonify({
                'success': True,
                'job_id': job_id,
                'message': 'Import job started in background',
                'status': 'pending',
                'use_background': True
            })
        
        # For small files or preview, process synchronously
        workbook = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
        
        # Get column mapping (already parsed above if background job path was taken)
        if 'column_map' not in locals():
            column_map = request.form.get('columnMap')
            if column_map:
                import json
                try:
                    column_map = json.loads(column_map)
                except:
                    column_map = {}
            else:
                column_map = {}
        
        # Get remaining options
        is_preview = request.form.get('preview', 'false').lower() == 'true'
        is_commit = request.form.get('commit', 'false').lower() == 'true'
        update_existing = request.form.get('updateExisting', 'false').lower() == 'true'
        force_commit = request.form.get('force', 'false').lower() == 'true'
        selected_sheets = request.form.get('selectedSheets')  # Comma-separated sheet names
        
        # Parse all sheets or selected sheets
        all_contacts = []
        sheet_names = workbook.sheetnames
        
        # Log available sheets for debugging
        logger.info(f"Excel file contains {len(sheet_names)} sheet(s): {', '.join(sheet_names)}")
        
        if selected_sheets:
            selected_sheet_list = [s.strip() for s in selected_sheets.split(',')]
            sheet_names = [s for s in sheet_names if s in selected_sheet_list]
            logger.info(f"Processing selected sheets: {', '.join(sheet_names)}")
        else:
            logger.info(f"Processing all sheets: {', '.join(sheet_names)}")
        
        for sheet_name in sheet_names:
            logger.info(f"Processing sheet: {sheet_name}")
            sheet = workbook[sheet_name]
            
            # Get headers from first row
            headers = []
            for cell in sheet[1]:
                header = str(cell.value).strip() if cell.value else ''
                # Apply column mapping if provided
                if column_map and header in column_map:
                    header = column_map[header]
                else:
                    # Auto-map common column names
                    header_lower = header.lower().replace(' ', '_').replace('-', '_')
                    if header_lower in ['name', 'full_name', 'contact_name']:
                        header = 'name'
                    elif header_lower in ['email', 'email_address', 'e_mail']:
                        header = 'email'
                    elif header_lower in ['phone', 'phone_number', 'mobile', 'tel']:
                        header = 'phone'
                    elif header_lower in ['company', 'company_name', 'organization']:
                        header = 'company'
                    elif header_lower in ['role', 'title', 'job_title', 'position', 'designation']:
                        header = 'role'
                    elif header_lower in ['linkedin', 'linkedin_url', 'linkedin_profile']:
                        header = 'linkedin'
                    elif header_lower in ['industry', 'sector']:
                        header = 'industry'
                    elif header_lower in ['city', 'location']:
                        header = 'city'
                    elif header_lower in ['lead_source', 'source', 'lead_source', 'source_name']:
                        header = 'lead_source'
                headers.append(header)
            
            # Parse rows
            row_count = 0
            skipped_count = 0
            for row in sheet.iter_rows(min_row=2, values_only=False):
                row_count += 1
                contact_data = {}
                for idx, cell in enumerate(row):
                    if idx < len(headers) and headers[idx] and cell.value:
                        contact_data[headers[idx]] = str(cell.value).strip()
                
                # Skip completely empty rows (no data at all)
                # Only skip if ALL of name, email, and phone are missing
                # This allows rows with just name, or just email, or just phone
                has_name = bool(contact_data.get('name'))
                has_email = bool(contact_data.get('email'))
                has_phone = bool(contact_data.get('phone'))
                
                # Skip only if row is completely empty (no name, email, or phone)
                if not has_name and not has_email and not has_phone:
                    skipped_count += 1
                    continue
                
                # Add sheet name
                contact_data['sheet'] = sheet_name
                
                all_contacts.append(contact_data)
            
            logger.info(f"Sheet '{sheet_name}': Processed {row_count} rows, imported {len([c for c in all_contacts if c.get('sheet') == sheet_name])} contacts, skipped {skipped_count} empty rows")
        
        if not all_contacts:
            return jsonify({'error': 'No contacts found in Excel file'}), 400
        
        # Normalize rows
        normalized_rows = []
        for row in all_contacts:
            normalized = {**row}
            
            # Extract domain from email if not provided
            if not normalized.get('domain') and normalized.get('email'):
                normalized['domain'] = resolve_best_domain(None, normalized['email'])
            
            # Ensure company name exists if we have domain but no company
            if not normalized.get('company') and not normalized.get('company_name') and normalized.get('domain'):
                # Use domain as company name if no company name provided
                domain_parts = normalized['domain'].split('.')
                if domain_parts:
                    normalized['company'] = domain_parts[0].title()
            
            # Normalize lead_source
            if normalized.get('leadSource') and not normalized.get('lead_source'):
                normalized['lead_source'] = normalized['leadSource']
            if not normalized.get('lead_source') and not normalized.get('leadSource') and normalized.get('sheet'):
                normalized['lead_source'] = normalized['sheet']
            
            normalized_rows.append(normalized)
        
        # Preview mode
        if is_preview:
            duplicates, uniques = find_duplicates(supabase, normalized_rows)
            
            sample = normalized_rows[:5]
            
            breakdown = {
                'emailMatches': len([d for d in duplicates if d.get('match_type') == 'email']),
                'phoneMatches': len([d for d in duplicates if d.get('match_type') == 'phone'])
            }
            
            # Format duplicates for display
            formatted_duplicates = []
            for dup in duplicates[:50]:  # Show up to 50 duplicates
                row = dup.get('row', {})
                match = dup.get('match', {})
                formatted_duplicates.append({
                    'name': row.get('name', ''),
                    'email': row.get('email', ''),
                    'phone': row.get('phone', ''),
                    'company': row.get('company', ''),
                    'match_type': dup.get('match_type', ''),
                    'existing_id': match.get('id'),
                    'existing_email': match.get('email', ''),
                    'existing_phone': match.get('phone', '')
                })
            
            # Format uniques for display
            formatted_uniques = []
            for unique in uniques[:50]:  # Show up to 50 uniques
                formatted_uniques.append({
                    'name': unique.get('name', ''),
                    'email': unique.get('email', ''),
                    'phone': unique.get('phone', ''),
                    'company': unique.get('company', ''),
                    'role': unique.get('role', ''),
                    'linkedin': unique.get('linkedin', '')
                })
            
            if not is_commit:
                return jsonify({
                    'preview': True,
                    'counts': {
                        'total': len(normalized_rows),
                        'duplicates': len(duplicates),
                        'uniques': len(uniques),
                        'toBeUpdated': len(duplicates),
                        'toBeInserted': len(uniques)
                    },
                    'breakdown': breakdown,
                    'duplicates': formatted_duplicates,
                    'uniques': formatted_uniques,
                    'sample': sample,
                    'sheets': sheet_names,  # Include all processed sheet names
                    'sheets_processed': len(sheet_names),
                    'sheets_total': len(workbook.sheetnames),
                    'duplicate_check': 'Checks for duplicates by: Email (primary) and Phone (secondary)'
                })
        
        # Commit mode
        if is_commit:
            if not is_preview and not force_commit:
                return jsonify({
                    'error': 'Commit requires a prior preview. Send preview=true to inspect duplicates/uniques, then commit=true to import. To override, send force=true.'
                }), 400
            
            # For large files, try to use background job as fallback
            if len(normalized_rows) > 500 or file_size > 1024 * 1024:
                logger.warning(f"Large file ({len(normalized_rows)} contacts, {file_size} bytes) being processed synchronously. Attempting background job fallback.")
                try:
                    # Re-read file content if needed
                    if 'file_content' not in locals():
                        file.seek(0)
                        file_content = file.read()
                    
                    # Create background job
                    import uuid
                    from datetime import datetime
                    from app.jobs.import_job import process_import_job
                    
                    job_id = str(uuid.uuid4())
                    
                    # Get import options
                    column_map = request.form.get('columnMap')
                    if column_map:
                        import json
                        try:
                            column_map = json.loads(column_map)
                        except:
                            column_map = {}
                    else:
                        column_map = {}
                    
                    options = {
                        'updateExisting': request.form.get('updateExisting', 'false').lower() == 'true',
                        'importOnlyNew': request.form.get('importOnlyNew', 'false').lower() == 'true',
                        'selectedSheets': request.form.get('selectedSheets', ''),
                        'columnMap': column_map
                    }
                    
                    # Create job record
                    job_data = {
                        'id': job_id,
                        'status': 'pending',
                        'file_name': file.filename,
                        'file_size': file_size,
                        'options': options,
                        'total_records': len(normalized_rows),
                        'processed_records': 0,
                        'imported_count': 0,
                        'error_count': 0,
                        'skipped_count': 0
                    }
                    
                    if user_id:
                        job_data['user_id'] = user_id
                    else:
                        supabase_user_id_fallback = session.get('user_id')
                        if supabase_user_id_fallback:
                            job_data['supabase_user_id'] = supabase_user_id_fallback
                    
                    insert_response = supabase.table('import_jobs').insert(job_data).execute()
                    if insert_response.data:
                        # Start background job
                        process_import_job(job_id, file_content, file.filename, column_map, options, current_app.app_context())
                        logger.info(f"Started background import job {job_id} for large file {file.filename}")
                        return jsonify({
                            'success': True,
                            'job_id': job_id,
                            'message': 'Import job started in background (large file detected)',
                            'status': 'pending',
                            'use_background': True
                        })
                except Exception as bg_error:
                    logger.error(f"Failed to create background job for large file: {bg_error}")
                    import traceback
                    logger.error(traceback.format_exc())
                    # Fall through to synchronous processing with warning
            
            # Check if user wants to import only new contacts (skip duplicates)
            import_only_new = request.form.get('importOnlyNew', 'false').lower() == 'true'
            
            if import_only_new:
                # Only import uniques, skip duplicates
                duplicates, uniques = find_duplicates(supabase, normalized_rows)
                logger.info(f"Import only new: {len(uniques)} new contacts, skipping {len(duplicates)} duplicates")
                rows_to_import = uniques
            else:
                # Import all (with updateExisting option)
                rows_to_import = normalized_rows
            
            from datetime import datetime
            import_start_time = datetime.now()
            logger.info(f"Starting import of {len(rows_to_import)} contacts (importOnlyNew={import_only_new})...")
            
            try:
                result = upsert_contacts(supabase, rows_to_import, {'updateExisting': update_existing})
            except Exception as import_error:
                logger.error(f"Error during import: {import_error}")
                import traceback
                logger.error(traceback.format_exc())
                return jsonify({
                    'error': f'Import failed: {str(import_error)}',
                    'hint': 'For large files (>500 contacts), the system should use background jobs. Please try again or contact support.'
                }), 500
            
            import_duration = (datetime.now() - import_start_time).total_seconds()
            logger.info(f"Import completed in {import_duration:.2f}s: {result.get('imported', 0)} imported, {len(result.get('errors', []))} errors")
            
            # Build detailed report with OK/Not OK status for each record
            detailed_report = []
            total_records = len(normalized_rows)
            imported_count = result.get('imported', 0)
            error_count = len(result.get('errors', []))
            
            # Map report entries by index
            report_map = {r.get('index', -1): r for r in result.get('report', [])}
            
            for idx, row in enumerate(normalized_rows):
                report_entry = report_map.get(idx, {})
                status = report_entry.get('status', 'skipped')
                error_msg = report_entry.get('error', '')
                
                if status == 'ok':
                    detailed_report.append({
                        'index': idx + 1,
                        'name': row.get('name', 'Unknown'),
                        'email': row.get('email', ''),
                        'status': 'OK',
                        'message': 'Imported successfully'
                    })
                elif status == 'skipped':
                    detailed_report.append({
                        'index': idx + 1,
                        'name': row.get('name', 'Unknown'),
                        'email': row.get('email', ''),
                        'status': 'Not OK',
                        'message': error_msg or 'Skipped - missing email and phone'
                    })
                else:
                    detailed_report.append({
                        'index': idx + 1,
                        'name': row.get('name', 'Unknown'),
                        'email': row.get('email', ''),
                        'status': 'Not OK',
                        'message': error_msg or 'Failed to import'
                    })
            
            if result.get('errors') and len(result['errors']) > 0:
                return jsonify({
                    'success': True,
                    'imported': imported_count,
                    'total': total_records,
                    'errors': result.get('errors', []),
                    'data': result.get('data', []),
                    'report': detailed_report,
                    'detailed_report': detailed_report,
                    'sheets': sheet_names
                }), 400
            
            return jsonify({
                'success': True,
                'imported': imported_count,
                'total': total_records,
                'data': result.get('data', []),
                'report': detailed_report,
                'detailed_report': detailed_report,
                'sheets': sheet_names
            })
        
        return jsonify({
            'error': 'Commit flag not provided. Send preview=true to inspect duplicates, then commit=true to import.'
        }), 400
        
    except Exception as e:
        logger.error(f"Error importing Excel file: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@contacts_bp.route('/api/contacts/export-excel', methods=['GET'])
@require_auth
@require_use_case('sheets_import_export')
def export_contacts_excel():
    """Export contacts to Excel file"""
    try:
        supabase = get_supabase_client(current_app)
        if not supabase:
            return jsonify({'error': 'Supabase not configured'}), 503
        
        # Get query parameters
        company_id = request.args.get('company_id')
        industry = request.args.get('industry')
        search = request.args.get('search')
        
        query = supabase.table('contacts').select('*')
        
        if company_id:
            query = query.eq('company_id', company_id)
        if industry:
            query = query.ilike('industry', f'%{industry}%')
        if search:
            query = query.or_(f'name.ilike.%{search}%,email.ilike.%{search}%,company.ilike.%{search}%')
        
        response = query.order('created_at', desc=True).execute()
        contacts = response.data or []
        
        if not contacts:
            return jsonify({'error': 'No contacts to export'}), 400
        
        # Create Excel workbook
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from flask import Response
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Contacts"
        
        # Headers
        headers = ['Name', 'Email', 'Phone', 'Company', 'Role', 'LinkedIn', 'Industry', 'City', 'Lead Source', 'Sheet', 'Created At']
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Add contact data
        for contact in contacts:
            row = [
                contact.get('name', ''),
                contact.get('email', ''),
                contact.get('phone', ''),
                contact.get('company', ''),
                contact.get('role', ''),
                contact.get('linkedin', ''),
                contact.get('industry', ''),
                contact.get('city', ''),
                contact.get('lead_source', ''),
                contact.get('sheet', ''),
                contact.get('created_at', '').strftime('%Y-%m-%d %H:%M:%S') if contact.get('created_at') else ''
            ]
            ws.append(row)
        
        # Auto-adjust column widths
        from openpyxl.utils import get_column_letter
        for idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(idx)
            max_length = max(
                len(str(header)),
                max([len(str(contact.get(header.lower().replace(' ', '_'), ''))) for contact in contacts] + [0])
            )
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
        
        # Save to BytesIO
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        from datetime import datetime
        filename = f"contacts_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return Response(
            output.read(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename={filename}'
            }
        )
        
    except Exception as e:
        logger.error(f"Error exporting contacts to Excel: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@contacts_bp.route('/api/contacts/export-sheets', methods=['POST'])
@require_auth
@require_use_case('sheets_import_export')
def export_contacts_sheets():
    """Export contacts to Google Sheets"""
    try:
        from app.integrations.google_sheets_client import GoogleSheetsClient
        
        supabase = get_supabase_client(current_app)
        if not supabase:
            return jsonify({'error': 'Supabase not configured'}), 503
        
        # Get query parameters
        data = request.get_json() or {}
        company_id = data.get('company_id') or request.args.get('company_id')
        industry = data.get('industry') or request.args.get('industry')
        search = data.get('search') or request.args.get('search')
        sheet_name = data.get('sheet_name', 'Contacts')
        
        query = supabase.table('contacts').select('*')
        
        if company_id:
            query = query.eq('company_id', company_id)
        if industry:
            query = query.ilike('industry', f'%{industry}%')
        if search:
            query = query.or_(f'name.ilike.%{search}%,email.ilike.%{search}%,company.ilike.%{search}%')
        
        response = query.order('created_at', desc=True).execute()
        contacts = response.data or []
        
        if not contacts:
            return jsonify({'error': 'No contacts to export'}), 400
        
        try:
            sheets_client = GoogleSheetsClient()
        except ValueError as e:
            return jsonify({
                'success': False,
                'error': 'Google Sheets credentials not found. Please configure GOOGLE_SHEETS_CREDENTIALS_PATH and GOOGLE_SHEETS_SPREADSHEET_ID in .env.local'
            }), 400
        
        # Prepare data for export
        headers = ['Name', 'Email', 'Phone', 'Company', 'Role', 'LinkedIn', 'Industry', 'City', 'Lead Source', 'Sheet', 'Created At']
        rows = [headers]
        
        for contact in contacts:
            row = [
                contact.get('name', ''),
                contact.get('email', ''),
                contact.get('phone', ''),
                contact.get('company', ''),
                contact.get('role', ''),
                contact.get('linkedin', ''),
                contact.get('industry', ''),
                contact.get('city', ''),
                contact.get('lead_source', ''),
                contact.get('sheet', ''),
                contact.get('created_at', '').strftime('%Y-%m-%d %H:%M:%S') if contact.get('created_at') else ''
            ]
            rows.append(row)
        
        # Export to Google Sheets
        try:
            spreadsheet = sheets_client.client.open_by_key(sheets_client.spreadsheet_id)
            
            # Try to get existing worksheet or create new
            try:
                worksheet = spreadsheet.worksheet(sheet_name)
                # Clear existing data
                worksheet.clear()
            except:
                worksheet = spreadsheet.add_worksheet(title=sheet_name, rows=len(rows) + 1, cols=len(headers))
            
            # Write data using batch update for better performance
            if len(rows) > 0:
                try:
                    worksheet.update('A1', rows, value_input_option='RAW')
                except Exception as update_error:
                    logger.warning(f"Error with update method, trying append: {update_error}")
                    # Fallback: clear and append row by row
                    worksheet.clear()
                    for row in rows:
                        worksheet.append_row(row)
            
            # Format header row (if gspread supports it)
            try:
                worksheet.format('A1:K1', {
                    'backgroundColor': {'red': 0.21, 'green': 0.38, 'blue': 0.57},
                    'textFormat': {'foregroundColor': {'red': 1.0, 'green': 1.0, 'blue': 1.0}, 'bold': True}
                })
            except Exception as format_error:
                logger.warning(f"Could not format header row: {format_error}")
            
            return jsonify({
                'success': True,
                'exported': len(contacts),
                'sheet_name': sheet_name,
                'message': f'Successfully exported {len(contacts)} contacts to Google Sheets'
            })
            
        except Exception as e:
            logger.error(f"Error writing to Google Sheets: {e}")
            return jsonify({
                'success': False,
                'error': f'Failed to export to Google Sheets: {str(e)}'
            }), 500
        
    except Exception as e:
        logger.error(f"Error exporting contacts to Google Sheets: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

