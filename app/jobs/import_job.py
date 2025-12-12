"""Background job processor for large contact imports"""
import logging
import threading
import openpyxl
from io import BytesIO
from datetime import datetime
from typing import Dict, Any, List, Optional
from app.services.contact_service import upsert_contacts, find_duplicates, resolve_best_domain
from app.supabase_client import get_supabase_client

logger = logging.getLogger(__name__)

# Global job registry to track running jobs
_active_jobs: Dict[str, threading.Thread] = {}


def process_import_job(
    job_id: str,
    file_content: bytes,
    file_name: str,
    column_map: Dict[str, str],
    options: Dict[str, Any],
    app_context
):
    """
    Process import job in background thread.
    This function runs in a separate thread to avoid blocking the main request.
    """
    def run_job():
        supabase = None
        try:
            # Get Flask app context
            with app_context:
                supabase = get_supabase_client(app_context.app)
                if not supabase:
                    _update_job_status(supabase, job_id, 'failed', progress_message='Supabase not configured')
                    return
                
                # Update job status to processing
                _update_job_status(supabase, job_id, 'processing', started_at=datetime.now(), progress_message='Reading Excel file...')
                
                # Read Excel file
                workbook = openpyxl.load_workbook(BytesIO(file_content), data_only=True, read_only=True)
                
                # Parse all sheets or selected sheets
                selected_sheets = options.get('selectedSheets')
                sheet_names = workbook.sheetnames
                if selected_sheets:
                    selected_sheet_list = [s.strip() for s in selected_sheets.split(',')]
                    sheet_names = [s for s in sheet_names if s in selected_sheet_list]
                
                # Collect all contacts
                all_contacts = []
                total_rows = 0
                
                for sheet_name in sheet_names:
                    sheet = workbook[sheet_name]
                    
                    # Get headers from first row
                    headers = []
                    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=False), None)
                    if not header_row:
                        continue
                    
                    for cell in header_row:
                        header = str(cell.value).strip() if cell.value else ''
                        # Apply column mapping if provided
                        if column_map and header in column_map:
                            header = column_map[header]
                        else:
                            # Auto-map common column names
                            header_lower = header.lower().replace(' ', '_').replace('-', '_')
                            if header_lower in ['name', 'full_name', 'contact_name']:
                                header = 'name'
                            elif header_lower in ['email', 'email_address', 'e_mail']:
                                header = 'email'
                            elif header_lower in ['phone', 'phone_number', 'mobile', 'tel']:
                                header = 'phone'
                            elif header_lower in ['company', 'company_name', 'organization']:
                                header = 'company'
                            elif header_lower in ['role', 'title', 'job_title', 'position', 'designation']:
                                header = 'role'
                            elif header_lower in ['linkedin', 'linkedin_url', 'linkedin_profile']:
                                header = 'linkedin'
                            elif header_lower in ['industry', 'sector']:
                                header = 'industry'
                            elif header_lower in ['city', 'location']:
                                header = 'city'
                            elif header_lower in ['lead_source', 'source', 'source_name']:
                                header = 'lead_source'
                        headers.append(header)
                    
                    # Parse rows using iterator for memory efficiency
                    _update_job_status(supabase, job_id, 'processing', progress_message=f'Parsing sheet: {sheet_name}...')
                    
                    row_count = 0
                    for row in sheet.iter_rows(min_row=2, values_only=False):
                        contact_data = {}
                        for idx, cell in enumerate(row):
                            if idx < len(headers) and headers[idx] and cell.value:
                                contact_data[headers[idx]] = str(cell.value).strip()
                        
                        # Skip empty rows
                        if not contact_data.get('name') and not contact_data.get('email') and not contact_data.get('phone'):
                            continue
                        
                        # Add sheet name
                        contact_data['sheet'] = sheet_name
                        all_contacts.append(contact_data)
                        row_count += 1
                        total_rows += 1
                        
                        # Update progress every 100 rows
                        if row_count % 100 == 0:
                            _update_job_status(
                                supabase, job_id, 'processing',
                                processed_records=len(all_contacts),
                                progress_message=f'Parsed {len(all_contacts)} contacts from {sheet_name}...'
                            )
                
                workbook.close()
                
                if not all_contacts:
                    _update_job_status(supabase, job_id, 'failed', progress_message='No contacts found in Excel file')
                    return
                
                # Normalize rows
                _update_job_status(supabase, job_id, 'processing', total_records=len(all_contacts), progress_message='Normalizing contact data...')
                normalized_rows = []
                for row in all_contacts:
                    normalized = {**row}
                    
                    # Extract domain from email if not provided
                    if not normalized.get('domain') and normalized.get('email'):
                        normalized['domain'] = resolve_best_domain(None, normalized['email'])
                    
                    # Normalize lead_source
                    if normalized.get('leadSource') and not normalized.get('lead_source'):
                        normalized['lead_source'] = normalized['leadSource']
                    if not normalized.get('lead_source') and not normalized.get('leadSource') and normalized.get('sheet'):
                        normalized['lead_source'] = normalized['sheet']
                    
                    normalized_rows.append(normalized)
                
                # Check for duplicates if importOnlyNew is enabled
                import_only_new = options.get('importOnlyNew', False)
                if import_only_new:
                    _update_job_status(supabase, job_id, 'processing', progress_message='Checking for duplicates...')
                    duplicates, uniques = find_duplicates(supabase, normalized_rows)
                    rows_to_import = uniques
                    skipped_count = len(duplicates)
                    logger.info(f"Import only new: {len(uniques)} new contacts, skipping {len(duplicates)} duplicates")
                else:
                    rows_to_import = normalized_rows
                    skipped_count = 0
                
                # Update total records
                _update_job_status(
                    supabase, job_id, 'processing',
                    total_records=len(rows_to_import),
                    skipped_count=skipped_count,
                    progress_message=f'Starting import of {len(rows_to_import)} contacts...'
                )
                
                # Process in batches with progress updates
                update_existing = options.get('updateExisting', False)
                batch_size = 100  # Process 100 contacts per batch
                total_batches = (len(rows_to_import) + batch_size - 1) // batch_size
                
                imported_count = 0
                error_count = 0
                errors = []
                
                for batch_idx in range(0, len(rows_to_import), batch_size):
                    batch = rows_to_import[batch_idx:batch_idx + batch_size]
                    batch_num = (batch_idx // batch_size) + 1
                    
                    _update_job_status(
                        supabase, job_id, 'processing',
                        processed_records=batch_idx,
                        progress_message=f'Processing batch {batch_num}/{total_batches} ({len(batch)} contacts)...'
                    )
                    
                    try:
                        result = upsert_contacts(supabase, batch, {'updateExisting': update_existing})
                        batch_imported = result.get('imported', 0)
                        batch_errors = result.get('errors', [])
                        
                        imported_count += batch_imported
                        error_count += len(batch_errors)
                        errors.extend(batch_errors)
                        
                    except Exception as batch_error:
                        logger.error(f"Error processing batch {batch_num}: {batch_error}")
                        error_count += len(batch)
                        errors.append({
                            'batch': batch_num,
                            'message': str(batch_error)
                        })
                
                # Final status update
                _update_job_status(
                    supabase, job_id, 'completed',
                    processed_records=len(rows_to_import),
                    imported_count=imported_count,
                    error_count=error_count,
                    completed_at=datetime.now(),
                    error_details=errors[:100],  # Limit to first 100 errors
                    progress_message=f'Import completed: {imported_count} imported, {error_count} errors, {skipped_count} skipped'
                )
                
                logger.info(f"Import job {job_id} completed: {imported_count} imported, {error_count} errors")
                
        except Exception as e:
            logger.error(f"Error processing import job {job_id}: {e}", exc_info=True)
            if supabase:
                _update_job_status(
                    supabase, job_id, 'failed',
                    error_count=1,
                    error_details=[{'message': str(e)}],
                    completed_at=datetime.now(),
                    progress_message=f'Import failed: {str(e)}'
                )
        finally:
            # Remove from active jobs
            if job_id in _active_jobs:
                del _active_jobs[job_id]
    
    # Start job in background thread
    thread = threading.Thread(target=run_job, daemon=True)
    thread.start()
    _active_jobs[job_id] = thread


def _update_job_status(
    supabase,
    job_id: str,
    status: str,
    total_records: Optional[int] = None,
    processed_records: Optional[int] = None,
    imported_count: Optional[int] = None,
    error_count: Optional[int] = None,
    skipped_count: Optional[int] = None,
    error_details: Optional[List] = None,
    started_at: Optional[datetime] = None,
    completed_at: Optional[datetime] = None,
    progress_message: Optional[str] = None
):
    """Update job status in database"""
    try:
        update_data = {'status': status}
        
        if total_records is not None:
            update_data['total_records'] = total_records
        if processed_records is not None:
            update_data['processed_records'] = processed_records
        if imported_count is not None:
            update_data['imported_count'] = imported_count
        if error_count is not None:
            update_data['error_count'] = error_count
        if skipped_count is not None:
            update_data['skipped_count'] = skipped_count
        if error_details is not None:
            import json
            update_data['error_details'] = json.dumps(error_details)
        if started_at:
            update_data['started_at'] = started_at.isoformat()
        if completed_at:
            update_data['completed_at'] = completed_at.isoformat()
        if progress_message:
            update_data['progress_message'] = progress_message
        
        supabase.table('import_jobs').update(update_data).eq('id', job_id).execute()
    except Exception as e:
        logger.error(f"Error updating job status for {job_id}: {e}")


def cancel_job(job_id: str):
    """Cancel a running import job"""
    if job_id in _active_jobs:
        # Mark as cancelled in database
        # Note: We can't actually stop the thread, but we can mark it
        try:
            from app.supabase_client import get_supabase_client
            from flask import current_app
            supabase = get_supabase_client(current_app)
            if supabase:
                supabase.table('import_jobs').update({
                    'status': 'cancelled',
                    'progress_message': 'Import cancelled by user'
                }).eq('id', job_id).execute()
        except:
            pass
        return True
    return False










